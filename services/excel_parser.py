"""
Excel/CSV parser for importing shot data
"""
import random
import string
from collections import OrderedDict
from pathlib import Path
from typing import Tuple

import pandas as pd
from loguru import logger
from openpyxl import Workbook


# Expected columns for the new format
EXPECTED_COLUMNS = ["shotId", "voiceRole", "voiceText", "scene", "characters", "imagePrompt", "videoPrompt"]


class ExcelParser:
    """Parser for Excel/CSV shot data"""

    def _generate_shot_id(self) -> str:
        """Generate a 6-character random ID for shot"""
        chars = string.ascii_lowercase + string.digits
        return ''.join(random.choices(chars, k=6))

    def parse(self, file_path: Path) -> Tuple[list, set, list]:
        """
        Parse Excel or CSV file and extract shots and characters

        Args:
            file_path: Path to the Excel/CSV file

        Returns:
            Tuple of (shots_list, characters_set, errors_list)
        """
        logger.info(f"Parsing file: {file_path}")
        errors = []

        try:
            # Read file based on extension
            if file_path.suffix.lower() == ".csv":
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)

            # Normalize column names
            df.columns = self._normalize_columns(df.columns.tolist())

            # Validate required columns
            missing_cols = self._check_required_columns(df.columns.tolist())
            if missing_cols:
                errors.append(f"Missing columns: {', '.join(missing_cols)}")
                logger.warning(f"Missing columns: {missing_cols}")

            # Group rows by shotId and merge dialogues
            shot_groups = OrderedDict()  # Preserve order
            all_characters = set()

            for idx, row in df.iterrows():
                try:
                    shot_id = str(row.get("shotId", "")) if pd.notna(row.get("shotId")) else ""
                    voice_role = str(row.get("voiceRole", "")) if pd.notna(row.get("voiceRole")) else ""
                    voice_text = str(row.get("voiceText", "")) if pd.notna(row.get("voiceText")) else ""
                    scene = str(row.get("scene", "")) if pd.notna(row.get("scene")) else ""
                    characters_str = str(row.get("characters", "")) if pd.notna(row.get("characters")) else ""
                    image_prompt = str(row.get("imagePrompt", "")) if pd.notna(row.get("imagePrompt")) else ""
                    video_prompt = str(row.get("videoPrompt", "")) if pd.notna(row.get("videoPrompt")) else ""

                    # Use row index as shotId if not provided
                    if not shot_id.strip():
                        shot_id = str(idx + 1)

                    # Parse characters from characters column
                    char_list = self._split_characters(characters_str)
                    for char in char_list:
                        all_characters.add(char)

                    # Also add voice role to characters if not empty
                    if voice_role.strip():
                        all_characters.add(voice_role.strip())

                    if shot_id not in shot_groups:
                        # First occurrence - create new shot
                        shot_groups[shot_id] = {
                            "dialogues": [],
                            "scene": scene,
                            "characters": char_list,
                            "imagePrompt": image_prompt,
                            "videoPrompt": video_prompt,
                        }

                    # Append dialogue if voice role and text exist
                    if voice_role.strip() or voice_text.strip():
                        shot_groups[shot_id]["dialogues"].append({
                            "role": voice_role.strip(),
                            "text": voice_text.strip(),
                        })

                except Exception as e:
                    error_msg = f"Row {idx + 2}: {str(e)}"
                    errors.append(error_msg)
                    logger.warning(error_msg)

            # Convert groups to shot list
            shots = []
            for seq, (shot_id, data) in enumerate(shot_groups.items(), 1):
                # Build script from dialogues
                script_lines = []
                for dialogue in data["dialogues"]:
                    if dialogue["role"] and dialogue["text"]:
                        script_lines.append(f"{dialogue['role']}: {dialogue['text']}")
                    elif dialogue["text"]:
                        script_lines.append(dialogue["text"])
                script = "\n".join(script_lines)

                shot = {
                    "id": self._generate_shot_id(),
                    "sequence": seq,
                    "scene": data["scene"],
                    "voiceActor": "",
                    "characters": data["characters"],
                    "emotion": "",
                    "intensity": "",
                    "script": script,
                    "imagePrompt": data["imagePrompt"],
                    "videoPrompt": data["videoPrompt"],
                    "images": [],
                    "selectedImageIndex": 0,
                    "videos": [],
                    "selectedVideoIndex": 0,
                    "videoUrl": "",
                    "audioUrl": "",
                    "status": "pending",
                }
                shots.append(shot)

            logger.info(f"Parsed {len(shots)} shots, {len(all_characters)} characters")
            return shots, all_characters, errors

        except Exception as e:
            logger.error(f"Failed to parse file: {e}")
            raise

    def _normalize_columns(self, columns: list) -> list:
        """Normalize column names to expected format"""
        normalized = []
        for col in columns:
            col_lower = str(col).lower().strip()

            # Try common variations and Chinese names
            mapping = {
                "shotId": ["镜头id", "镜头", "shotid", "shot_id", "id", "shot"],
                "voiceRole": ["配音角色", "voicerole", "voice_role", "role", "角色", "peiyin_juese"],
                "voiceText": ["配音内容", "voicetext", "voice_text", "text", "内容", "对白", "台词", "peiyin_neirong"],
                "scene": ["场景", "scene", "changjing"],
                "characters": ["出场角色", "characters", "chuchang_juese", "chuchangjuese", "出场"],
                "imagePrompt": ["图片提示词", "imageprompt", "tupian_tishici", "image", "tupiantishici", "image_prompt", "图片"],
                "videoPrompt": ["视频提示词", "videoprompt", "shipin_tishici", "video", "shipintishici", "video_prompt", "视频"],
            }

            found = False
            for key, variants in mapping.items():
                if col_lower in [v.lower() for v in variants] or col in variants:
                    normalized.append(key)
                    found = True
                    break

            if not found:
                normalized.append(col_lower)

        return normalized

    def _check_required_columns(self, columns: list) -> list:
        """Check for missing required columns"""
        required = ["imagePrompt"]
        return [col for col in required if col not in columns]

    def _split_characters(self, char_str: str) -> list:
        """Split character string into list"""
        if not char_str or char_str == "nan":
            return []
        # Split by common separators
        for sep in [",", "/", "|", "、"]:
            if sep in char_str:
                return [c.strip() for c in char_str.split(sep) if c.strip()]
        return [char_str.strip()] if char_str.strip() else []

    def export_template(self, file_path: Path):
        """Export an empty Excel template"""
        wb = Workbook()
        ws = wb.active
        ws.title = "镜头列表"

        # Headers (Chinese)
        headers = ["镜头ID", "配音角色", "配音内容", "场景", "出场角色", "图片提示词", "视频提示词"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Example rows - demonstrating merge behavior
        examples = [
            ["1", "旁白", "在一个宁静的小镇上，住着一位年轻的画家。", "小镇街道", "画家", "宁静的小镇街道，阳光洒落", "镜头缓慢推进"],
            ["2", "画家", "今天的天气真好啊。", "画室内", "画家", "画家站在窗前，望向窗外", "特写画家的表情"],
            ["2", "旁白", "他望向窗外，若有所思。", "", "", "", ""],  # Same shotId=2, will be merged
            ["3", "画家", "我要出门走走。", "画室门口", "画家", "画家推开门，准备出门", "跟随镜头"],
        ]
        for row_idx, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Set column widths for better readability
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 30
        ws.column_dimensions["G"].width = 25

        wb.save(file_path)
        logger.info(f"Template exported to: {file_path}")
